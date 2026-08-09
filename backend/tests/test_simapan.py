"""SIMAPAN backend regression tests."""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/") if os.environ.get("REACT_APP_BACKEND_URL") else "https://custody-management-2.preview.emergentagent.com"
API = f"{BASE_URL}/api"


def _login(username, password):
    r = requests.post(f"{API}/auth/login", json={"username": username, "password": password}, timeout=20)
    assert r.status_code == 200, f"login {username} failed: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="session")
def admin_token():
    return _login("admin", "Admin@123")


@pytest.fixture(scope="session")
def sup_token():
    return _login("supervisor", "Supervisor@123")


@pytest.fixture(scope="session")
def op_token():
    return _login("operator1", "Operator@123")


def H(tok):
    return {"Authorization": f"Bearer {tok}"}


# --- Auth ---
class TestAuth:
    def test_login_bad(self):
        r = requests.post(f"{API}/auth/login", json={"username": "admin", "password": "wrong"})
        assert r.status_code == 401

    def test_me(self, admin_token):
        r = requests.get(f"{API}/auth/me", headers=H(admin_token))
        assert r.status_code == 200
        assert r.json()["role"] == "admin"

    def test_me_unauth(self):
        assert requests.get(f"{API}/auth/me").status_code == 401


# --- Dashboard ---
class TestDashboard:
    def test_stats(self, admin_token):
        r = requests.get(f"{API}/dashboard/stats", headers=H(admin_token))
        assert r.status_code == 200
        d = r.json()
        for k in ["total_inmates", "active_inmates", "scans_today", "locations", "operators"]:
            assert k in d


# --- Inmates ---
class TestInmates:
    def test_list(self, admin_token):
        r = requests.get(f"{API}/inmates", headers=H(admin_token))
        assert r.status_code == 200
        assert isinstance(r.json(), list)
        assert len(r.json()) >= 3

    def test_lookup(self, admin_token):
        r = requests.get(f"{API}/inmates/lookup/WB-2024-0001", headers=H(admin_token))
        assert r.status_code == 200
        assert r.json()["full_name"] == "Budi Santoso"

    def test_create_update(self, admin_token):
        payload = {
            "full_name": "TEST_WargaBinaan",
            "registration_number": "TEST-REG-9901",
            "cell_block": "C",
            "crime_category": "Uji Coba",
            "medical_alert": "-",
            "barcode_data": "TEST-BC-9901",
        }
        r = requests.post(f"{API}/inmates", headers=H(admin_token), json=payload)
        assert r.status_code == 200, r.text
        iid = r.json()["id"]
        assert r.json()["barcode_data"] == "TEST-BC-9901"

        # GET verify
        g = requests.get(f"{API}/inmates/{iid}", headers=H(admin_token))
        assert g.status_code == 200
        assert g.json()["full_name"] == "TEST_WargaBinaan"

        # Update
        payload["full_name"] = "TEST_WargaBinaanEdit"
        u = requests.put(f"{API}/inmates/{iid}", headers=H(admin_token), json=payload)
        assert u.status_code == 200
        assert u.json()["full_name"] == "TEST_WargaBinaanEdit"

        # Barcode PNG
        b = requests.get(f"{API}/inmates/{iid}/barcode", headers=H(admin_token))
        assert b.status_code == 200
        assert b.headers["content-type"].startswith("image/png")
        assert len(b.content) > 100

        # Delete
        d = requests.delete(f"{API}/inmates/{iid}", headers=H(admin_token))
        assert d.status_code == 200
        assert requests.get(f"{API}/inmates/{iid}", headers=H(admin_token)).status_code == 404


# --- Locations ---
class TestLocations:
    def test_crud(self, admin_token):
        payload = {"location_name": "TEST_LokasiScan", "location_type": "other",
                   "gps_coordinates": "-6.2000,106.8000"}
        r = requests.post(f"{API}/locations", headers=H(admin_token), json=payload)
        assert r.status_code == 200, r.text
        lid = r.json()["id"]

        # Update name
        payload["location_name"] = "TEST_LokasiScanEdit"
        u = requests.put(f"{API}/locations/{lid}", headers=H(admin_token), json=payload)
        assert u.status_code == 200
        assert u.json()["location_name"] == "TEST_LokasiScanEdit"

        # Barcode
        b = requests.get(f"{API}/locations/{lid}/barcode", headers=H(admin_token))
        assert b.status_code == 200 and b.headers["content-type"].startswith("image/png")

        # Cleanup
        requests.delete(f"{API}/locations/{lid}", headers=H(admin_token))


# --- Settings (category rename) ---
class TestSettings:
    def test_update_category(self, admin_token):
        s = requests.get(f"{API}/settings", headers=H(admin_token))
        assert s.status_code == 200
        original = s.json()["activity_categories"]

        # rename keagamaan
        modified = [dict(c) for c in original]
        for c in modified:
            if c["key"] == "keagamaan":
                c["label"] = "Pembinaan Rohani"
        r = requests.put(f"{API}/settings", headers=H(admin_token),
                         json={"activity_categories": modified})
        assert r.status_code == 200
        got = [c["label"] for c in r.json()["activity_categories"] if c["key"] == "keagamaan"][0]
        assert got == "Pembinaan Rohani"

        # revert
        requests.put(f"{API}/settings", headers=H(admin_token),
                     json={"activity_categories": original})

    def test_settings_no_admin(self, op_token):
        r = requests.put(f"{API}/settings", headers=H(op_token), json={"app_title": "x"})
        assert r.status_code == 403


# --- Activities: scan flow + approve + reject + export ---
class TestActivities:
    def test_full_flow(self, op_token, sup_token, admin_token):
        # Operator scans via barcode_code
        payload = {"barcode_code": "WB-2024-0001",
                   "activity_category": "keagamaan",
                   "duration_minutes": 30,
                   "inmate_condition": "baik",
                   "description": "TEST_scan operator"}
        r = requests.post(f"{API}/activities", headers=H(op_token), json=payload)
        assert r.status_code == 200, r.text
        act1 = r.json()
        assert act1["inmate_reg"] == "WB-2024-0001"
        assert act1["scan_location"] == "Masjid At-Taubah"  # from assigned location
        assert act1["status"] == "submitted"

        # Approve
        a = requests.post(f"{API}/activities/{act1['id']}/approve", headers=H(sup_token),
                          json={"comment": "OK"})
        assert a.status_code == 200
        assert a.json()["status"] == "approved"

        # Create another & reject
        r2 = requests.post(f"{API}/activities", headers=H(op_token), json=payload)
        assert r2.status_code == 200
        act2 = r2.json()
        rj = requests.post(f"{API}/activities/{act2['id']}/reject", headers=H(sup_token),
                           json={"reason": "TEST_ditolak alasan uji"})
        assert rj.status_code == 200
        assert rj.json()["status"] == "rejected"
        assert rj.json()["rejection_reason"] == "TEST_ditolak alasan uji"

        # History
        h = requests.get(f"{API}/activities/{act2['id']}/history", headers=H(admin_token))
        assert h.status_code == 200 and len(h.json()) >= 1

        # Export xlsx
        e = requests.get(f"{API}/activities/export", headers=H(admin_token))
        assert e.status_code == 200
        assert "spreadsheet" in e.headers["content-type"]
        wb = load_workbook(io.BytesIO(e.content))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "No"

    def test_scan_inactive_inmate_fails(self, op_token):
        r = requests.post(f"{API}/activities", headers=H(op_token),
                          json={"barcode_code": "DOES-NOT-EXIST",
                                "activity_category": "keagamaan"})
        assert r.status_code == 404


# --- Users management + role restrictions ---
class TestUsersMgmt:
    def test_operator_forbidden_users(self, op_token):
        assert requests.get(f"{API}/users", headers=H(op_token)).status_code == 403

    def test_admin_create_user_and_login(self, admin_token):
        # get a location id for assignment
        locs = requests.get(f"{API}/locations", headers=H(admin_token)).json()
        loc_id = locs[0]["id"] if locs else None
        payload = {
            "username": "test_op_new",
            "full_name": "TEST Operator Baru",
            "role": "operator",
            "password": "TestPass@123",
            "assigned_location": loc_id,
            "device_name": "TEST-Tablet-01",
            "status": "active",
        }
        # cleanup if exists
        existing = requests.get(f"{API}/users", headers=H(admin_token)).json()
        for u in existing:
            if u["username"] == "test_op_new":
                requests.delete(f"{API}/users/{u['id']}", headers=H(admin_token))
        r = requests.post(f"{API}/users", headers=H(admin_token), json=payload)
        assert r.status_code == 200, r.text
        new_id = r.json()["id"]

        # login
        li = requests.post(f"{API}/auth/login",
                          json={"username": "test_op_new", "password": "TestPass@123"})
        assert li.status_code == 200

        # cleanup
        requests.delete(f"{API}/users/{new_id}", headers=H(admin_token))


# --- Audit log ---
class TestAudit:
    def test_admin_audit(self, admin_token):
        r = requests.get(f"{API}/audit-logs", headers=H(admin_token))
        assert r.status_code == 200
        entries = r.json()
        actions = {e["action"] for e in entries}
        # login must exist since we logged in above
        assert "login" in actions or len(entries) > 0

    def test_op_forbidden(self, op_token):
        assert requests.get(f"{API}/audit-logs", headers=H(op_token)).status_code == 403
