import os
import io
import json
import uuid
import logging
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import Optional, List

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env")

import bcrypt
import jwt
import qrcode
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage, TextDelta, StreamDone
except ImportError:
    LlmChat = UserMessage = TextDelta = StreamDone = None

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI(title="KAWAN PAS API")
api_router = APIRouter(prefix="/api")

JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGORITHM = "HS256"
ACCESS_MINUTES = 60 * 12

logger = logging.getLogger(__name__)


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def new_id():
    return str(uuid.uuid4())


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_access_token(user_id: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_MINUTES),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def clean_user(u: dict) -> dict:
    u = dict(u)
    u.pop("_id", None)
    u.pop("hashed_password", None)
    return u


async def get_current_user(request: Request) -> dict:
    token = None
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
    if not token:
        token = request.cookies.get("access_token")
    if not token:
        token = request.query_params.get("token")
    if not token:
        raise HTTPException(status_code=401, detail="Tidak terautentikasi")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token kedaluwarsa")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token tidak valid")
    user = await db.users.find_one({"id": payload["sub"]})
    if not user or user.get("status") != "active":
        raise HTTPException(status_code=401, detail="Pengguna tidak ditemukan atau nonaktif")
    await db.users.update_one({"id": user["id"]}, {"$set": {"last_activity": now_iso()}})
    return clean_user(user)


def require_roles(*roles):
    async def dep(user: dict = Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Akses ditolak")
        return user
    return dep


def _strip_ids(obj):
    if isinstance(obj, dict):
        return {k: _strip_ids(v) for k, v in obj.items() if k != "_id"}
    if isinstance(obj, list):
        return [_strip_ids(v) for v in obj]
    try:
        from bson import ObjectId
        if isinstance(obj, ObjectId):
            return str(obj)
    except Exception:
        pass
    return obj


async def log_audit(user: dict, entity_type: str, entity_id: str, action: str,
                    changes: Optional[dict] = None, request: Optional[Request] = None):
    doc = {
        "id": new_id(),
        "user_id": user["id"],
        "username": user.get("username"),
        "entity_type": entity_type,
        "entity_id": entity_id,
        "action": action,
        "changes_json": _strip_ids(changes or {}),
        "ip_address": request.client.host if request and request.client else None,
        "device_info": request.headers.get("user-agent") if request else None,
        "timestamp": now_iso(),
    }
    await db.audit_log.insert_one(doc)


# ---------------- Settings ----------------

DEFAULT_SETTINGS = {
    "key": "app_settings",
    "app_title": "KAWAN PAS",
    "app_subtitle": "Sistem Monitoring Aktivitas Warga Binaan",
    "institution_name": "Lembaga Pemasyarakatan",
    "activity_categories": [
        {"key": "ibadah", "label": "Ibadah", "module": "pembinaan"},
        {"key": "kerja_bengkel", "label": "Kerja Bengkel", "module": "pembinaan"},
        {"key": "keagamaan", "label": "Pembinaan Keagamaan", "module": "pembinaan"},
        {"key": "keterampilan", "label": "Pembinaan Keterampilan", "module": "pembinaan"},
        {"key": "kesehatan", "label": "Layanan Kesehatan", "module": "pembinaan"},
        {"key": "pendidikan", "label": "Pendidikan & Pelatihan", "module": "pembinaan"},
        {"key": "olahraga", "label": "Olahraga & Rekreasi", "module": "pembinaan"},
        {"key": "keamanan", "label": "Titik Keamanan", "module": "keamanan"},
        {"key": "lainnya", "label": "Lainnya", "module": "pembinaan"},
    ],
    "location_types": [
        {"key": "religious", "label": "Keagamaan", "module": "pembinaan"},
        {"key": "skills", "label": "Keterampilan", "module": "pembinaan"},
        {"key": "health", "label": "Kesehatan", "module": "pembinaan"},
        {"key": "education", "label": "Pendidikan", "module": "pembinaan"},
        {"key": "sports", "label": "Olahraga", "module": "pembinaan"},
        {"key": "security", "label": "Keamanan", "module": "keamanan"},
        {"key": "other", "label": "Lainnya", "module": "pembinaan"},
    ],
    "inmate_conditions": [
        {"key": "baik", "label": "Baik"},
        {"key": "sakit", "label": "Sakit"},
        {"key": "perlu_perhatian", "label": "Perlu Perhatian"},
    ],
}


async def get_settings_doc() -> dict:
    s = await db.settings.find_one({"key": "app_settings"})
    if not s:
        s = dict(DEFAULT_SETTINGS)
        await db.settings.insert_one(s)
    s.pop("_id", None)
    return s


async def migrate_settings_modules():
    """Migrasi non-destruktif: tandai module pembinaan/keamanan pada kategori & tipe lokasi
    yang sudah ada di DB, dan tambahkan kategori baru (ibadah, kerja_bengkel) bila belum ada."""
    s = await get_settings_doc()
    default_cat_by_key = {c["key"]: c for c in DEFAULT_SETTINGS["activity_categories"]}
    default_loc_by_key = {c["key"]: c for c in DEFAULT_SETTINGS["location_types"]}
    changed = False

    cats = s.get("activity_categories", [])
    existing_cat_keys = {c["key"] for c in cats}
    for c in cats:
        if not c.get("module"):
            c["module"] = default_cat_by_key.get(c["key"], {}).get("module", "pembinaan")
            changed = True
    for key, dc in default_cat_by_key.items():
        if key not in existing_cat_keys:
            cats.append(dict(dc))
            changed = True

    loc_types = s.get("location_types", [])
    existing_loc_keys = {c["key"] for c in loc_types}
    for c in loc_types:
        if not c.get("module"):
            c["module"] = default_loc_by_key.get(c["key"], {}).get("module", "pembinaan")
            changed = True
    for key, dc in default_loc_by_key.items():
        if key not in existing_loc_keys:
            loc_types.append(dict(dc))
            changed = True

    if changed:
        await db.settings.update_one(
            {"key": "app_settings"},
            {"$set": {"activity_categories": cats, "location_types": loc_types}})


@api_router.get("/settings")
async def get_settings(user: dict = Depends(get_current_user)):
    return await get_settings_doc()


@api_router.put("/settings")
async def update_settings(payload: dict, request: Request,
                          user: dict = Depends(require_roles("admin"))):
    allowed = ["app_title", "app_subtitle", "institution_name",
               "activity_categories", "location_types", "inmate_conditions"]
    before = await get_settings_doc()
    update = {k: payload[k] for k in allowed if k in payload}
    await db.settings.update_one({"key": "app_settings"}, {"$set": update}, upsert=True)
    after = await get_settings_doc()
    await log_audit(user, "settings", "app_settings", "update",
                    {"before": before, "after": after}, request)
    return after


# ---------------- Auth ----------------

class LoginPayload(BaseModel):
    username: str
    password: str
    device_info: Optional[str] = None


@api_router.post("/auth/login")
async def login(payload: LoginPayload, response: Response, request: Request):
    user = await db.users.find_one({"username": payload.username.strip().lower()})
    if not user or not verify_password(payload.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Username atau password salah")
    if user.get("status") != "active":
        raise HTTPException(status_code=403, detail="Akun nonaktif. Hubungi admin.")
    token = create_access_token(user["id"], user["role"])
    await db.users.update_one({"id": user["id"]}, {"$set": {
        "last_login": now_iso(),
        "last_activity": now_iso(),
        "last_device_info": payload.device_info or request.headers.get("user-agent"),
    }})
    response.set_cookie(key="access_token", value=token, httponly=True,
                        samesite="lax", max_age=ACCESS_MINUTES * 60, path="/")
    await log_audit(clean_user(user), "users", user["id"], "login", {}, request)
    return {"token": token, "user": clean_user(user)}


@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


# ---------------- Users ----------------

class UserPayload(BaseModel):
    username: str
    full_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    position: Optional[str] = None
    assigned_location: Optional[str] = None
    device_name: Optional[str] = None
    role: str = "operator"
    password: Optional[str] = None
    status: str = "active"
    start_date: Optional[str] = None
    end_date: Optional[str] = None


@api_router.get("/users")
async def list_users(user: dict = Depends(require_roles("admin"))):
    users = await db.users.find().sort("created_at", -1).to_list(500)
    return [clean_user(u) for u in users]


@api_router.post("/users")
async def create_user(payload: UserPayload, request: Request,
                      user: dict = Depends(require_roles("admin"))):
    username = payload.username.strip().lower()
    if await db.users.find_one({"username": username}):
        raise HTTPException(status_code=400, detail="Username sudah digunakan")
    if not payload.password:
        raise HTTPException(status_code=400, detail="Password wajib diisi")
    doc = payload.model_dump()
    doc.pop("password")
    doc.update({
        "id": new_id(),
        "username": username,
        "hashed_password": hash_password(payload.password),
        "last_login": None,
        "last_activity": None,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    })
    await db.users.insert_one(doc)
    await log_audit(user, "users", doc["id"], "create", {"after": clean_user(doc)}, request)
    return clean_user(doc)


@api_router.put("/users/{user_id}")
async def update_user(user_id: str, payload: UserPayload, request: Request,
                      user: dict = Depends(require_roles("admin"))):
    existing = await db.users.find_one({"id": user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")
    username = payload.username.strip().lower()
    clash = await db.users.find_one({"username": username, "id": {"$ne": user_id}})
    if clash:
        raise HTTPException(status_code=400, detail="Username sudah digunakan")
    update = payload.model_dump()
    pw = update.pop("password", None)
    if pw:
        update["hashed_password"] = hash_password(pw)
    update["username"] = username
    update["updated_at"] = now_iso()
    await db.users.update_one({"id": user_id}, {"$set": update})
    after = await db.users.find_one({"id": user_id})
    await log_audit(user, "users", user_id, "update",
                    {"before": clean_user(existing), "after": clean_user(after)}, request)
    return clean_user(after)


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request,
                      user: dict = Depends(require_roles("admin"))):
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Tidak dapat menghapus akun sendiri")
    existing = await db.users.find_one({"id": user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")
    await db.users.delete_one({"id": user_id})
    await log_audit(user, "users", user_id, "delete", {"before": clean_user(existing)}, request)
    return {"ok": True}


# ---------------- Inmates ----------------

class InmatePayload(BaseModel):
    full_name: str
    registration_number: Optional[str] = None
    identity_number: Optional[str] = None
    photo_url: Optional[str] = None
    status: str = "active"
    date_entry: Optional[str] = None
    estimated_release_date: Optional[str] = None
    cell_block: Optional[str] = None
    crime_category: Optional[str] = None
    medical_alert: Optional[str] = None
    barcode_data: Optional[str] = None
    age: Optional[int] = None
    religion: Optional[str] = None
    mp_1_3: Optional[str] = None
    mp_1_2: Optional[str] = None
    mp_2_3: Optional[str] = None
    program_notes: Optional[str] = None
    # Rekam medis detail
    blood_type: Optional[str] = None
    height_cm: Optional[int] = None
    weight_kg: Optional[int] = None
    allergies: Optional[str] = None
    chronic_conditions: Optional[str] = None
    current_medications: Optional[str] = None
    medical_notes: Optional[str] = None


@api_router.get("/inmates")
async def list_inmates(search: Optional[str] = None, status: Optional[str] = None,
                       cell_block: Optional[str] = None,
                       user: dict = Depends(get_current_user)):
    q = {}
    if status:
        q["status"] = status
    if cell_block:
        q["cell_block"] = cell_block
    if search:
        q["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"registration_number": {"$regex": search, "$options": "i"}},
            {"identity_number": {"$regex": search, "$options": "i"}},
        ]
    inmates = await db.inmates.find(q, {"_id": 0}).sort("full_name", 1).to_list(2000)
    return inmates


@api_router.post("/inmates")
async def create_inmate(payload: InmatePayload, request: Request,
                        user: dict = Depends(require_roles("admin", "supervisor"))):
    reg = (payload.registration_number or "").strip()
    if not reg:
        count = await db.inmates.count_documents({})
        reg = f"WB-{datetime.now().year}-{count + 1:04d}"
    if await db.inmates.find_one({"registration_number": reg}):
        raise HTTPException(status_code=400, detail="Nomor registrasi sudah digunakan")
    doc = payload.model_dump()
    doc["registration_number"] = reg
    if not doc.get("barcode_data"):
        doc["barcode_data"] = reg
    doc.update({"id": new_id(), "created_at": now_iso(), "updated_at": now_iso()})
    await db.inmates.insert_one(doc)
    doc.pop("_id", None)
    await log_audit(user, "inmates", doc["id"], "create", {"after": doc}, request)
    return doc


@api_router.get("/inmates/lookup/{code}")
async def lookup_inmate(code: str, user: dict = Depends(get_current_user)):
    inmate = await db.inmates.find_one(
        {"$or": [{"barcode_data": code}, {"registration_number": code}]}, {"_id": 0})
    if not inmate:
        raise HTTPException(status_code=404, detail="Warga binaan tidak ditemukan")
    return inmate


@api_router.get("/inmates/{inmate_id}")
async def get_inmate(inmate_id: str, user: dict = Depends(get_current_user)):
    inmate = await db.inmates.find_one({"id": inmate_id}, {"_id": 0})
    if not inmate:
        raise HTTPException(status_code=404, detail="Warga binaan tidak ditemukan")
    return inmate


@api_router.put("/inmates/{inmate_id}")
async def update_inmate(inmate_id: str, payload: InmatePayload, request: Request,
                        user: dict = Depends(require_roles("admin", "supervisor"))):
    existing = await db.inmates.find_one({"id": inmate_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Warga binaan tidak ditemukan")
    update = payload.model_dump(exclude_unset=True)
    if update.get("registration_number"):
        clash = await db.inmates.find_one({
            "registration_number": update["registration_number"],
            "id": {"$ne": inmate_id}})
        if clash:
            raise HTTPException(status_code=400, detail="Nomor registrasi sudah digunakan")
    if "barcode_data" in update and not update["barcode_data"]:
        update["barcode_data"] = update.get("registration_number") or existing["registration_number"]
    update["updated_at"] = now_iso()
    await db.inmates.update_one({"id": inmate_id}, {"$set": update})
    after = await db.inmates.find_one({"id": inmate_id}, {"_id": 0})
    existing.pop("_id", None)
    await log_audit(user, "inmates", inmate_id, "update",
                    {"before": existing, "after": after}, request)
    return after


@api_router.delete("/inmates/{inmate_id}")
async def delete_inmate(inmate_id: str, request: Request,
                        user: dict = Depends(require_roles("admin"))):
    existing = await db.inmates.find_one({"id": inmate_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Warga binaan tidak ditemukan")
    await db.inmates.delete_one({"id": inmate_id})
    await log_audit(user, "inmates", inmate_id, "delete", {"before": existing}, request)
    return {"ok": True}


def make_qr_png(data: str) -> bytes:
    img = qrcode.make(data, box_size=10, border=2)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


@api_router.get("/inmates/{inmate_id}/barcode")
async def inmate_barcode(inmate_id: str, download: bool = False,
                         user: dict = Depends(get_current_user)):
    inmate = await db.inmates.find_one({"id": inmate_id})
    if not inmate:
        raise HTTPException(status_code=404, detail="Warga binaan tidak ditemukan")
    png = make_qr_png(inmate.get("barcode_data") or inmate["registration_number"])
    headers = {}
    if download:
        fname = f"barcode_{inmate['registration_number']}.png"
        headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return StreamingResponse(io.BytesIO(png), media_type="image/png", headers=headers)


# ---------------- Kartu Identitas (ukuran kartu ATM/CR80) ----------------

CARD_W_MM = 85.6
CARD_H_MM = 54.0


def _fetch_photo_reader(photo_url: Optional[str]):
    if not photo_url:
        return None
    try:
        from reportlab.lib.utils import ImageReader
        if photo_url.startswith("data:image"):
            import base64 as _b64
            header, b64data = photo_url.split(",", 1)
            return ImageReader(io.BytesIO(_b64.b64decode(b64data)))
        if photo_url.startswith("http://") or photo_url.startswith("https://"):
            import urllib.request
            req = urllib.request.Request(photo_url, headers={"User-Agent": "KAWAN-PAS/1.0"})
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = resp.read()
            return ImageReader(io.BytesIO(data))
    except Exception:
        return None
    return None


def _draw_id_card(c, x_mm, y_mm, inmate: dict, institution_name: str, app_title: str):
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    x, y = x_mm * mm, y_mm * mm
    w, h = CARD_W_MM * mm, CARD_H_MM * mm

    c.saveState()
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(1)
    c.rect(x, y, w, h, stroke=1, fill=0)

    # Header strip
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.rect(x, y + h - 8 * mm, w, 8 * mm, stroke=0, fill=1)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 6.5)
    c.drawString(x + 3 * mm, y + h - 5.3 * mm, institution_name.upper()[:42])
    c.setFont("Helvetica", 5)
    c.drawString(x + 3 * mm, y + h - 7.3 * mm, "KARTU IDENTITAS WARGA BINAAN")

    # Photo box
    photo_x, photo_y = x + 3 * mm, y + 3 * mm
    photo_w, photo_h = 20 * mm, h - 13 * mm
    c.setFillColorRGB(0.92, 0.92, 0.92)
    c.rect(photo_x, photo_y, photo_w, photo_h, stroke=1, fill=1)
    photo_reader = _fetch_photo_reader(inmate.get("photo_url"))
    if photo_reader is not None:
        try:
            c.drawImage(photo_reader, photo_x, photo_y, width=photo_w, height=photo_h,
                       preserveAspectRatio=True, mask="auto")
        except Exception:
            photo_reader = None
    if photo_reader is None:
        initials = "".join([p[0] for p in (inmate.get("full_name") or "?").split()[:2]]).upper()
        c.setFillColorRGB(0.5, 0.5, 0.5)
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(photo_x + photo_w / 2, photo_y + photo_h / 2 - 3, initials or "?")

    # Text block
    text_x = photo_x + photo_w + 3 * mm
    text_top = y + h - 11 * mm
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 8.5)
    c.drawString(text_x, text_top, (inmate.get("full_name") or "-")[:26])
    c.setFont("Helvetica", 6)
    rows = [
        ("No. Register", inmate.get("registration_number") or "-"),
        ("Blok", inmate.get("cell_block") or "-"),
        ("Gol. Darah", inmate.get("blood_type") or "-"),
        ("Agama", inmate.get("religion") or "-"),
    ]
    ty = text_top - 4.2 * mm
    for label, val in rows:
        c.setFont("Helvetica", 5.5)
        c.drawString(text_x, ty, f"{label}:")
        c.setFont("Helvetica-Bold", 5.5)
        c.drawString(text_x + 15 * mm, ty, str(val)[:18])
        ty -= 3.4 * mm

    if inmate.get("medical_alert"):
        c.setFillColorRGB(0.6, 0, 0)
        c.setFont("Helvetica-Bold", 5)
        c.drawString(text_x, y + 3.5 * mm, ("PERINGATAN MEDIS: " + inmate["medical_alert"])[:34])

    # QR code
    qr_size = 16 * mm
    qr_x = x + w - qr_size - 3 * mm
    qr_y = y + 3 * mm
    png = make_qr_png(inmate.get("barcode_data") or inmate["registration_number"])
    c.drawImage(ImageReader(io.BytesIO(png)), qr_x, qr_y, width=qr_size, height=qr_size)
    c.setFont("Helvetica", 4.5)
    c.drawCentredString(qr_x + qr_size / 2, qr_y - 2.2, app_title[:20])

    c.restoreState()


@api_router.get("/inmates/{inmate_id}/card")
async def inmate_card(inmate_id: str, user: dict = Depends(get_current_user)):
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.lib.units import mm
    inmate = await db.inmates.find_one({"id": inmate_id}, {"_id": 0})
    if not inmate:
        raise HTTPException(status_code=404, detail="Warga binaan tidak ditemukan")
    settings = await get_settings_doc()
    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=(CARD_W_MM * mm, CARD_H_MM * mm))
    _draw_id_card(c, 0, 0, inmate, settings.get("institution_name", ""), settings.get("app_title", "KAWAN PAS"))
    c.showPage()
    c.save()
    buf.seek(0)
    fname = f"kartu_{inmate['registration_number']}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@api_router.get("/inmates/cards/batch")
async def inmates_cards_batch(status: Optional[str] = "active",
                              user: dict = Depends(require_roles("admin", "supervisor"))):
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    q = {}
    if status:
        q["status"] = status
    inmates = await db.inmates.find(q, {"_id": 0}).sort("full_name", 1).to_list(2000)
    if not inmates:
        raise HTTPException(status_code=404, detail="Tidak ada warga binaan untuk dicetak")
    settings = await get_settings_doc()
    inst = settings.get("institution_name", "")
    title = settings.get("app_title", "KAWAN PAS")

    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=A4)
    page_w, page_h = A4
    margin = 10 * mm
    gap_x, gap_y = 4 * mm, 4 * mm
    cols = int((page_w - 2 * margin + gap_x) // (CARD_W_MM * mm + gap_x))
    cols = max(cols, 1)
    rows = int((page_h - 2 * margin + gap_y) // (CARD_H_MM * mm + gap_y))
    rows = max(rows, 1)
    per_page = cols * rows

    for idx, inmate in enumerate(inmates):
        pos = idx % per_page
        if idx > 0 and pos == 0:
            c.showPage()
        col = pos % cols
        row = pos // cols
        x_mm = (margin + col * (CARD_W_MM * mm + gap_x)) / mm
        y_mm = (page_h - margin - (row + 1) * CARD_H_MM * mm - row * gap_y) / mm
        _draw_id_card(c, x_mm, y_mm, inmate, inst, title)
    c.showPage()
    c.save()
    buf.seek(0)
    fname = f"kartu_warga_binaan_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------- Locations ----------------

class LocationPayload(BaseModel):
    location_name: str
    location_type: str = "other"
    gps_coordinates: Optional[str] = None
    description: Optional[str] = None


@api_router.get("/locations")
async def list_locations(user: dict = Depends(get_current_user)):
    return await db.locations.find({}, {"_id": 0}).sort("location_name", 1).to_list(500)


@api_router.post("/locations")
async def create_location(payload: LocationPayload, request: Request,
                          user: dict = Depends(require_roles("admin"))):
    if await db.locations.find_one({"location_name": payload.location_name.strip()}):
        raise HTTPException(status_code=400, detail="Nama lokasi sudah digunakan")
    doc = payload.model_dump()
    doc["location_name"] = doc["location_name"].strip()
    doc.update({"id": new_id(), "created_at": now_iso()})
    await db.locations.insert_one(doc)
    doc.pop("_id", None)
    await log_audit(user, "locations", doc["id"], "create", {"after": doc}, request)
    return doc


@api_router.put("/locations/{location_id}")
async def update_location(location_id: str, payload: LocationPayload, request: Request,
                          user: dict = Depends(require_roles("admin"))):
    existing = await db.locations.find_one({"id": location_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lokasi tidak ditemukan")
    update = payload.model_dump()
    update["location_name"] = update["location_name"].strip()
    clash = await db.locations.find_one({
        "location_name": update["location_name"], "id": {"$ne": location_id}})
    if clash:
        raise HTTPException(status_code=400, detail="Nama lokasi sudah digunakan")
    await db.locations.update_one({"id": location_id}, {"$set": update})
    after = await db.locations.find_one({"id": location_id}, {"_id": 0})
    await log_audit(user, "locations", location_id, "update",
                    {"before": existing, "after": after}, request)
    return after


@api_router.delete("/locations/{location_id}")
async def delete_location(location_id: str, request: Request,
                          user: dict = Depends(require_roles("admin"))):
    existing = await db.locations.find_one({"id": location_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lokasi tidak ditemukan")
    await db.locations.delete_one({"id": location_id})
    await log_audit(user, "locations", location_id, "delete", {"before": existing}, request)
    return {"ok": True}


@api_router.get("/locations/{location_id}/barcode")
async def location_barcode(location_id: str, download: bool = False,
                           user: dict = Depends(get_current_user)):
    loc = await db.locations.find_one({"id": location_id})
    if not loc:
        raise HTTPException(status_code=404, detail="Lokasi tidak ditemukan")
    png = make_qr_png(f"LOC:{location_id}")
    headers = {}
    if download:
        fname = f"barcode_lokasi_{loc['location_name'].replace(' ', '_')}.png"
        headers["Content-Disposition"] = f'attachment; filename="{fname}"'
    return StreamingResponse(io.BytesIO(png), media_type="image/png", headers=headers)


# ---------------- Keamanan: Lalu Lintas Warga Binaan ----------------

class SecurityCrossingPayload(BaseModel):
    inmate_id: Optional[str] = None
    barcode_code: Optional[str] = None
    location_id: Optional[str] = None
    direction: str = "keluar"  # masuk | keluar
    purpose: Optional[str] = None
    escort_officer: Optional[str] = None
    notes: Optional[str] = None
    scan_timestamp: Optional[str] = None


DIRECTION_LABEL = {"masuk": "Masuk", "keluar": "Keluar"}


@api_router.get("/security/crossings")
async def list_crossings(location_id: Optional[str] = None, inmate_id: Optional[str] = None,
                         direction: Optional[str] = None,
                         date_from: Optional[str] = None, date_to: Optional[str] = None,
                         user: dict = Depends(get_current_user)):
    q = {}
    if location_id:
        q["location_id"] = location_id
    if inmate_id:
        q["inmate_id"] = inmate_id
    if direction:
        q["direction"] = direction
    if user["role"] == "operator":
        q["operator_user_id"] = user["id"]
    if date_from or date_to:
        q["scan_timestamp"] = {}
        if date_from:
            q["scan_timestamp"]["$gte"] = date_from
        if date_to:
            q["scan_timestamp"]["$lte"] = (date_to + "T23:59:59+00:00") if len(date_to) == 10 else date_to
    items = await db.security_crossings.find(q, {"_id": 0}).sort("scan_timestamp", -1).to_list(5000)
    return items


@api_router.post("/security/crossings")
async def create_crossing(payload: SecurityCrossingPayload, request: Request,
                          user: dict = Depends(require_roles("admin", "supervisor", "operator"))):
    inmate = None
    if payload.inmate_id:
        inmate = await db.inmates.find_one({"id": payload.inmate_id})
    elif payload.barcode_code:
        inmate = await db.inmates.find_one({"$or": [
            {"barcode_data": payload.barcode_code},
            {"registration_number": payload.barcode_code}]})
    if not inmate:
        raise HTTPException(status_code=404, detail="Warga binaan tidak ditemukan dari kode tersebut")

    location_name = None
    if payload.location_id:
        loc = await db.locations.find_one({"id": payload.location_id})
        if loc:
            location_name = loc["location_name"]
    if not location_name and user.get("assigned_location"):
        loc = await db.locations.find_one({"id": user["assigned_location"]})
        if loc:
            location_name = loc["location_name"]
            payload.location_id = loc["id"]

    direction = payload.direction if payload.direction in ("masuk", "keluar") else "keluar"
    doc = {
        "id": new_id(),
        "inmate_id": inmate["id"],
        "inmate_name": inmate["full_name"],
        "inmate_reg": inmate["registration_number"],
        "cell_block": inmate.get("cell_block"),
        "operator_user_id": user["id"],
        "operator_name": user["full_name"],
        "scan_timestamp": payload.scan_timestamp or now_iso(),
        "checkpoint_location": location_name,
        "location_id": payload.location_id,
        "direction": direction,
        "direction_label": DIRECTION_LABEL.get(direction, direction),
        "purpose": payload.purpose,
        "escort_officer": payload.escort_officer,
        "notes": payload.notes,
        "device_info": request.headers.get("user-agent"),
        "created_at": now_iso(),
    }
    await db.security_crossings.insert_one(doc)
    await log_audit(user, "security_crossings", doc["id"], "create", {"after": doc}, request)
    doc.pop("_id", None)
    return doc


@api_router.get("/security/crossings/export")
async def export_crossings(location_id: Optional[str] = None, direction: Optional[str] = None,
                           date_from: Optional[str] = None, date_to: Optional[str] = None,
                           request: Request = None, user: dict = Depends(get_current_user)):
    q = {}
    if location_id:
        q["location_id"] = location_id
    if direction:
        q["direction"] = direction
    if user["role"] == "operator":
        q["operator_user_id"] = user["id"]
    if date_from or date_to:
        q["scan_timestamp"] = {}
        if date_from:
            q["scan_timestamp"]["$gte"] = date_from
        if date_to:
            q["scan_timestamp"]["$lte"] = (date_to + "T23:59:59+00:00") if len(date_to) == 10 else date_to
    items = await db.security_crossings.find(q, {"_id": 0}).sort("scan_timestamp", -1).to_list(20000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lalu Lintas"
    headers = ["No", "Waktu", "No. Registrasi", "Nama Warga Binaan", "Blok", "Titik Keamanan",
               "Arah", "Tujuan/Keperluan", "Petugas Pengawal", "Catatan", "Operator"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A0A0A")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 4)
    for i, a in enumerate(items, 1):
        ws.append([
            i, a.get("scan_timestamp"), a.get("inmate_reg"), a.get("inmate_name"), a.get("cell_block"),
            a.get("checkpoint_location"), a.get("direction_label") or a.get("direction"),
            a.get("purpose"), a.get("escort_officer"), a.get("notes"), a.get("operator_name"),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"lalu_lintas_keamanan_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await log_audit(user, "security_crossings", "-", "export",
                    {"filters": {"location_id": location_id, "direction": direction,
                                 "date_from": date_from, "date_to": date_to}, "count": len(items)}, request)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@api_router.get("/security/crossings/report")
async def crossings_report(location_id: Optional[str] = None, direction: Optional[str] = None,
                           date_from: Optional[str] = None, date_to: Optional[str] = None,
                           request: Request = None, user: dict = Depends(get_current_user)):
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    q = {}
    if location_id:
        q["location_id"] = location_id
    if direction:
        q["direction"] = direction
    if user["role"] == "operator":
        q["operator_user_id"] = user["id"]
    if date_from or date_to:
        q["scan_timestamp"] = {}
        if date_from:
            q["scan_timestamp"]["$gte"] = date_from
        if date_to:
            q["scan_timestamp"]["$lte"] = (date_to + "T23:59:59+00:00") if len(date_to) == 10 else date_to
    items = await db.security_crossings.find(q, {"_id": 0}).sort("scan_timestamp", 1).to_list(20000)
    settings = await get_settings_doc()

    buf = io.BytesIO()
    page = landscape(A4)
    c = pdfcanvas.Canvas(buf, pagesize=page)
    w, h = page
    margin = 12 * mm

    def header():
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(w / 2, h - margin, settings.get("institution_name", ""))
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(w / 2, h - margin - 5.5 * mm, "LAPORAN LALU LINTAS WARGA BINAAN")
        c.setFont("Helvetica", 8)
        rng = f"{date_from or '-'} s.d. {date_to or '-'}"
        c.drawCentredString(w / 2, h - margin - 10 * mm, f"Periode: {rng}  |  Dicetak: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
        c.line(margin, h - margin - 13 * mm, w - margin, h - margin - 13 * mm)

    col_widths = [10 * mm, 28 * mm, 22 * mm, 40 * mm, 15 * mm, 38 * mm, 16 * mm, 45 * mm, 35 * mm]
    col_titles = ["No", "Waktu", "No. Reg", "Nama", "Blok", "Titik Keamanan", "Arah", "Tujuan", "Petugas Pengawal"]

    def table_header(y):
        c.setFont("Helvetica-Bold", 7.5)
        cx = margin
        c.setFillColorRGB(0.9, 0.9, 0.9)
        c.rect(margin, y - 5 * mm, sum(col_widths), 5.5 * mm, stroke=1, fill=1)
        c.setFillColorRGB(0, 0, 0)
        for i, t in enumerate(col_titles):
            c.drawString(cx + 1.5 * mm, y - 3.5 * mm, t)
            cx += col_widths[i]
        return y - 6 * mm

    header()
    y = h - margin - 17 * mm
    y = table_header(y)
    c.setFont("Helvetica", 7)
    row_h = 5 * mm

    for i, a in enumerate(items, 1):
        if y < margin + 20 * mm:
            c.showPage()
            header()
            y = h - margin - 17 * mm
            y = table_header(y)
            c.setFont("Helvetica", 7)
        ts = a.get("scan_timestamp", "")
        try:
            ts_disp = datetime.fromisoformat(ts.replace("Z", "+00:00")).strftime("%d/%m %H:%M")
        except Exception:
            ts_disp = (ts or "-")[:16]
        vals = [str(i), ts_disp, a.get("inmate_reg") or "-", (a.get("inmate_name") or "-")[:28],
                a.get("cell_block") or "-", (a.get("checkpoint_location") or "-")[:26],
                a.get("direction_label") or "-", (a.get("purpose") or "-")[:32],
                (a.get("escort_officer") or "-")[:24]]
        cx = margin
        for j, v in enumerate(vals):
            c.drawString(cx + 1.5 * mm, y - 3.5 * mm, v)
            cx += col_widths[j]
        c.line(margin, y - row_h, margin + sum(col_widths), y - row_h)
        y -= row_h

    y -= 12 * mm
    if y < margin + 20 * mm:
        c.showPage()
        y = h - margin - 20 * mm
    c.setFont("Helvetica", 8)
    c.drawString(w - 75 * mm, y, f"{settings.get('institution_name','')}, {datetime.now().strftime('%d %B %Y')}")
    c.drawString(w - 75 * mm, y - 5 * mm, "Kepala Regu Keamanan,")
    c.line(w - 75 * mm, y - 22 * mm, w - 20 * mm, y - 22 * mm)
    c.drawString(w - 75 * mm, y - 26 * mm, "( ......................................... )")

    c.showPage()
    c.save()
    buf.seek(0)
    fname = f"laporan_lalulintas_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    if request is not None:
        await log_audit(user, "security_crossings", "-", "export",
                        {"type": "report_pdf", "filters": {"location_id": location_id, "direction": direction,
                                                            "date_from": date_from, "date_to": date_to},
                         "count": len(items)}, request)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@api_router.get("/security/crossings/{crossing_id}")
async def get_crossing(crossing_id: str, user: dict = Depends(get_current_user)):
    a = await db.security_crossings.find_one({"id": crossing_id}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Data lalu lintas tidak ditemukan")
    return a


@api_router.delete("/security/crossings/{crossing_id}")
async def delete_crossing(crossing_id: str, request: Request,
                          user: dict = Depends(require_roles("admin"))):
    existing = await db.security_crossings.find_one({"id": crossing_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Data lalu lintas tidak ditemukan")
    await db.security_crossings.delete_one({"id": crossing_id})
    await log_audit(user, "security_crossings", crossing_id, "delete", {"before": existing}, request)
    return {"ok": True}


def _wrap_lines(c, text, x, y, max_width, font="Helvetica", size=8, leading=10):
    from reportlab.pdfbase.pdfmetrics import stringWidth
    words = (text or "-").split()
    line = ""
    for w in words:
        trial = f"{line} {w}".strip()
        if stringWidth(trial, font, size) > max_width and line:
            c.drawString(x, y, line)
            y -= leading
            line = w
        else:
            line = trial
    if line:
        c.drawString(x, y, line)
        y -= leading
    return y


@api_router.get("/security/crossings/{crossing_id}/bon")
async def crossing_bon(crossing_id: str, user: dict = Depends(get_current_user)):
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.lib.pagesizes import A6
    from reportlab.lib.units import mm
    a = await db.security_crossings.find_one({"id": crossing_id}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Data lalu lintas tidak ditemukan")
    settings = await get_settings_doc()
    buf = io.BytesIO()
    w, h = A6
    c = pdfcanvas.Canvas(buf, pagesize=A6)

    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(w / 2, h - 14 * mm, settings.get("institution_name", "")[:40])
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(w / 2, h - 19 * mm, "BON LALU LINTAS WARGA BINAAN")
    c.setLineWidth(0.7)
    c.line(8 * mm, h - 22 * mm, w - 8 * mm, h - 22 * mm)

    ts = a.get("scan_timestamp", "")
    try:
        ts_disp = datetime.fromisoformat(ts.replace("Z", "+00:00")).strftime("%d-%m-%Y %H:%M")
    except Exception:
        ts_disp = ts

    rows = [
        ("No. Bon", a["id"][:8].upper()),
        ("Waktu", ts_disp),
        ("Nama", a.get("inmate_name") or "-"),
        ("No. Register", a.get("inmate_reg") or "-"),
        ("Blok", a.get("cell_block") or "-"),
        ("Titik Keamanan", a.get("checkpoint_location") or "-"),
        ("Arah", a.get("direction_label") or a.get("direction") or "-"),
        ("Petugas Pengawal", a.get("escort_officer") or "-"),
        ("Operator", a.get("operator_name") or "-"),
    ]
    y = h - 28 * mm
    c.setFont("Helvetica", 8.5)
    for label, val in rows:
        c.setFont("Helvetica-Bold", 8)
        c.drawString(8 * mm, y, f"{label}")
        c.setFont("Helvetica", 8)
        c.drawString(38 * mm, y, str(val)[:26])
        y -= 6 * mm

    c.setFont("Helvetica-Bold", 8)
    c.drawString(8 * mm, y, "Tujuan/Keperluan")
    y -= 5 * mm
    c.setFont("Helvetica", 8)
    y = _wrap_lines(c, a.get("purpose") or "-", 8 * mm, y, w - 16 * mm, size=8, leading=4.5 * mm)
    if a.get("notes"):
        y -= 2 * mm
        c.setFont("Helvetica-Bold", 8)
        c.drawString(8 * mm, y, "Catatan")
        y -= 5 * mm
        c.setFont("Helvetica", 8)
        y = _wrap_lines(c, a.get("notes"), 8 * mm, y, w - 16 * mm, size=8, leading=4.5 * mm)

    y -= 8 * mm
    c.setFont("Helvetica", 8)
    c.drawString(8 * mm, y, "Petugas Keamanan,")
    c.drawString(w - 55 * mm, y, "Mengetahui,")
    y -= 20 * mm
    c.line(8 * mm, y, 45 * mm, y)
    c.line(w - 55 * mm, y, w - 8 * mm, y)
    y -= 4 * mm
    c.setFont("Helvetica", 7)
    c.drawString(8 * mm, y, "( ......................... )")
    c.drawString(w - 55 * mm, y, "( ......................... )")

    c.showPage()
    c.save()
    buf.seek(0)
    fname = f"bon_lalulintas_{a.get('inmate_reg','')}_{a['id'][:8]}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------- Activities ----------------

class ActivityPayload(BaseModel):
    inmate_id: Optional[str] = None
    barcode_code: Optional[str] = None
    location_id: Optional[str] = None
    scan_location: Optional[str] = None
    activity_category: Optional[str] = None
    duration_minutes: Optional[int] = None
    description: Optional[str] = None
    inmate_condition: str = "baik"
    scan_timestamp: Optional[str] = None
    status: str = "submitted"
    photos: List[str] = []


async def activity_out(a: dict) -> dict:
    a.pop("_id", None)
    return a


@api_router.get("/activities")
async def list_activities(status: Optional[str] = None, category: Optional[str] = None,
                          location_id: Optional[str] = None, inmate_id: Optional[str] = None,
                          date_from: Optional[str] = None, date_to: Optional[str] = None,
                          user: dict = Depends(get_current_user)):
    q = {}
    if status:
        q["status"] = status
    if category:
        q["activity_category"] = category
    if location_id:
        q["location_id"] = location_id
    if inmate_id:
        q["inmate_id"] = inmate_id
    if user["role"] == "operator":
        q["operator_user_id"] = user["id"]
    if date_from or date_to:
        q["scan_timestamp"] = {}
        if date_from:
            q["scan_timestamp"]["$gte"] = date_from
        if date_to:
            q["scan_timestamp"]["$lte"] = (date_to + "T23:59:59+00:00") if len(date_to) == 10 else date_to
    items = await db.activities.find(q, {"_id": 0}).sort("scan_timestamp", -1).to_list(5000)
    return items


@api_router.post("/activities")
async def create_activity(payload: ActivityPayload, request: Request,
                          user: dict = Depends(require_roles("admin", "supervisor", "operator"))):
    inmate = None
    if payload.inmate_id:
        inmate = await db.inmates.find_one({"id": payload.inmate_id})
    elif payload.barcode_code:
        inmate = await db.inmates.find_one({"$or": [
            {"barcode_data": payload.barcode_code},
            {"registration_number": payload.barcode_code}]})
    if not inmate:
        raise HTTPException(status_code=404, detail="Warga binaan tidak ditemukan dari kode tersebut")
    if inmate.get("status") != "active":
        raise HTTPException(status_code=400, detail=f"Warga binaan berstatus {inmate.get('status')}, tidak dapat discan")

    location_name = payload.scan_location
    if payload.location_id:
        loc = await db.locations.find_one({"id": payload.location_id})
        if loc:
            location_name = loc["location_name"]
    if not location_name and user.get("assigned_location"):
        loc = await db.locations.find_one({"id": user["assigned_location"]})
        if loc:
            location_name = loc["location_name"]
            payload.location_id = loc["id"]

    settings = await get_settings_doc()
    cat_label = next((c["label"] for c in settings["activity_categories"]
                      if c["key"] == payload.activity_category), payload.activity_category)

    doc = {
        "id": new_id(),
        "inmate_id": inmate["id"],
        "inmate_name": inmate["full_name"],
        "inmate_reg": inmate["registration_number"],
        "operator_user_id": user["id"],
        "operator_name": user["full_name"],
        "scan_timestamp": payload.scan_timestamp or now_iso(),
        "scan_location": location_name,
        "location_id": payload.location_id,
        "activity_category": payload.activity_category,
        "activity_category_label": cat_label,
        "duration_minutes": payload.duration_minutes,
        "description": payload.description,
        "inmate_condition": payload.inmate_condition,
        "photos": payload.photos or [],
        "status": payload.status if payload.status in ("draft", "submitted") else "submitted",
        "approval_user_id": None,
        "approval_user_name": None,
        "approval_timestamp": None,
        "rejection_reason": None,
        "device_info": request.headers.get("user-agent"),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.activities.insert_one(doc)
    await log_audit(user, "activities", doc["id"], "create",
                    {"after": {k: v for k, v in doc.items()}}, request)
    doc.pop("_id", None)
    return doc


@api_router.get("/activities/export")
async def export_activities(status: Optional[str] = None, category: Optional[str] = None,
                            location_id: Optional[str] = None,
                            date_from: Optional[str] = None, date_to: Optional[str] = None,
                            request: Request = None,
                            user: dict = Depends(get_current_user)):
    q = {}
    if status:
        q["status"] = status
    if category:
        q["activity_category"] = category
    if location_id:
        q["location_id"] = location_id
    if user["role"] == "operator":
        q["operator_user_id"] = user["id"]
    if date_from or date_to:
        q["scan_timestamp"] = {}
        if date_from:
            q["scan_timestamp"]["$gte"] = date_from
        if date_to:
            q["scan_timestamp"]["$lte"] = (date_to + "T23:59:59+00:00") if len(date_to) == 10 else date_to
    items = await db.activities.find(q, {"_id": 0}).sort("scan_timestamp", -1).to_list(20000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Aktivitas"
    headers = ["No", "Waktu Scan", "No. Registrasi", "Nama Warga Binaan", "Lokasi",
               "Kategori Aktivitas", "Durasi (menit)", "Kondisi", "Status",
               "Operator", "Disetujui Oleh", "Waktu Persetujuan", "Alasan Penolakan", "Deskripsi"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A0A0A")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 4)
    status_map = {"draft": "Draf", "submitted": "Menunggu", "approved": "Disetujui", "rejected": "Ditolak"}
    cond_map = {"baik": "Baik", "sakit": "Sakit", "perlu_perhatian": "Perlu Perhatian"}
    for i, a in enumerate(items, 1):
        ws.append([
            i, a.get("scan_timestamp"), a.get("inmate_reg"), a.get("inmate_name"),
            a.get("scan_location"), a.get("activity_category_label") or a.get("activity_category"),
            a.get("duration_minutes"), cond_map.get(a.get("inmate_condition"), a.get("inmate_condition")),
            status_map.get(a.get("status"), a.get("status")), a.get("operator_name"),
            a.get("approval_user_name"), a.get("approval_timestamp"),
            a.get("rejection_reason"), a.get("description"),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"laporan_aktivitas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await log_audit(user, "activities", "-", "export",
                    {"filters": {"status": status, "category": category,
                                 "location_id": location_id, "date_from": date_from,
                                 "date_to": date_to}, "count": len(items)}, request)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@api_router.get("/activities/{activity_id}")
async def get_activity(activity_id: str, user: dict = Depends(get_current_user)):
    a = await db.activities.find_one({"id": activity_id}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")
    return a


@api_router.get("/activities/{activity_id}/history")
async def activity_history(activity_id: str, user: dict = Depends(get_current_user)):
    items = await db.approvals.find({"activity_id": activity_id}, {"_id": 0}).sort("timestamp", 1).to_list(100)
    return items


async def change_status(activity_id: str, new_status: str, user: dict, request: Request,
                        reason: Optional[str] = None):
    a = await db.activities.find_one({"id": activity_id})
    if not a:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")
    old_status = a["status"]
    update = {"status": new_status, "updated_at": now_iso()}
    if new_status == "approved":
        update.update({"approval_user_id": user["id"], "approval_user_name": user["full_name"],
                       "approval_timestamp": now_iso(), "rejection_reason": None})
    elif new_status == "rejected":
        update.update({"approval_user_id": user["id"], "approval_user_name": user["full_name"],
                       "approval_timestamp": now_iso(), "rejection_reason": reason})
    await db.activities.update_one({"id": activity_id}, {"$set": update})
    await db.approvals.insert_one({
        "id": new_id(), "activity_id": activity_id, "old_status": old_status,
        "new_status": new_status, "approved_by_user_id": user["id"],
        "approved_by_name": user["full_name"], "reason": reason, "timestamp": now_iso(),
    })
    await log_audit(user, "activities", activity_id,
                    "approve" if new_status == "approved" else new_status,
                    {"before": {"status": old_status}, "after": {"status": new_status},
                     "reason": reason}, request)
    return await db.activities.find_one({"id": activity_id}, {"_id": 0})


class ApprovePayload(BaseModel):
    comment: Optional[str] = None


class RejectPayload(BaseModel):
    reason: str


@api_router.post("/activities/{activity_id}/approve")
async def approve_activity(activity_id: str, payload: ApprovePayload, request: Request,
                           user: dict = Depends(require_roles("admin", "supervisor"))):
    return await change_status(activity_id, "approved", user, request, payload.comment)


@api_router.post("/activities/{activity_id}/reject")
async def reject_activity(activity_id: str, payload: RejectPayload, request: Request,
                          user: dict = Depends(require_roles("admin", "supervisor"))):
    if not payload.reason.strip():
        raise HTTPException(status_code=400, detail="Alasan penolakan wajib diisi")
    return await change_status(activity_id, "rejected", user, request, payload.reason)


@api_router.post("/activities/{activity_id}/submit")
async def submit_activity(activity_id: str, request: Request,
                          user: dict = Depends(get_current_user)):
    a = await db.activities.find_one({"id": activity_id})
    if not a:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")
    if user["role"] == "operator" and a["operator_user_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    return await change_status(activity_id, "submitted", user, request)


@api_router.delete("/activities/{activity_id}")
async def delete_activity(activity_id: str, request: Request,
                          user: dict = Depends(require_roles("admin"))):
    a = await db.activities.find_one({"id": activity_id}, {"_id": 0})
    if not a:
        raise HTTPException(status_code=404, detail="Aktivitas tidak ditemukan")
    await db.activities.delete_one({"id": activity_id})
    await log_audit(user, "activities", activity_id, "delete", {"before": a}, request)
    return {"ok": True}


# ---------------- Audit ----------------

@api_router.get("/audit-logs")
async def list_audit(entity_type: Optional[str] = None, action: Optional[str] = None,
                     limit: int = 300, user: dict = Depends(require_roles("admin"))):
    q = {}
    if entity_type:
        q["entity_type"] = entity_type
    if action:
        q["action"] = action
    items = await db.audit_log.find(q, {"_id": 0}).sort("timestamp", -1).to_list(min(limit, 1000))
    return [_strip_ids(i) for i in items]


# ---------------- Dashboard ----------------

@api_router.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date().isoformat()
    total_inmates = await db.inmates.count_documents({})
    active_inmates = await db.inmates.count_documents({"status": "active"})
    scans_today = await db.activities.count_documents({"scan_timestamp": {"$gte": today}})
    total_scans = await db.activities.count_documents({})
    pending = await db.activities.count_documents({"status": "submitted"})
    locations = await db.locations.count_documents({})
    operators = await db.users.count_documents({"role": "operator", "status": "active"})

    cat_pipeline = [
        {"$match": {"scan_timestamp": {"$gte": today}}},
        {"$group": {"_id": "$activity_category_label", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    by_category = await db.activities.aggregate(cat_pipeline).to_list(20)
    loc_pipeline = [
        {"$match": {"scan_timestamp": {"$gte": today}}},
        {"$group": {"_id": "$scan_location", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    by_location = await db.activities.aggregate(loc_pipeline).to_list(20)
    recent = await db.activities.find({}, {"_id": 0}).sort("scan_timestamp", -1).to_list(8)

    inmate_statuses = {}
    for s in ["active", "leave", "released", "transferred"]:
        inmate_statuses[s] = await db.inmates.count_documents({"status": s})

    return {
        "total_inmates": total_inmates,
        "active_inmates": active_inmates,
        "scans_today": scans_today,
        "total_scans": total_scans,
        "pending_approvals": pending,
        "locations": locations,
        "operators": operators,
        "by_category_today": [{"label": c["_id"] or "Lainnya", "count": c["count"]} for c in by_category],
        "by_location_today": [{"label": l["_id"] or "-", "count": l["count"]} for l in by_location],
        "recent_activities": recent,
        "inmate_statuses": inmate_statuses,
    }


@api_router.get("/")
async def root():
    return {"message": "KAWAN PAS API aktif"}


# ---------------- AI Assistant (Claude) ----------------

EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY")
CLAUDE_PROVIDER, CLAUDE_MODEL = "anthropic", "claude-sonnet-4-6"

AI_SYSTEM = (
    "Kamu adalah asisten AI KAWAN PAS, sistem monitoring aktivitas warga binaan "
    "di lembaga pemasyarakatan Indonesia. Jawab SELALU dalam Bahasa Indonesia yang formal dan ringkas. "
    "Bantu admin dan supervisor memahami data aktivitas, warga binaan, lokasi pemindaian, dan persetujuan. "
    "Jawaban harus berdasarkan data sistem yang diberikan. Jika data tidak tersedia, katakan dengan jujur."
)


async def build_ai_context() -> str:
    today = datetime.now(timezone.utc).date().isoformat()
    lines = []
    lines.append(f"Tanggal hari ini: {today}")
    for label, count in [
        ("Total warga binaan", await db.inmates.count_documents({})),
        ("Warga binaan aktif", await db.inmates.count_documents({"status": "active"})),
        ("Pemindaian hari ini", await db.activities.count_documents({"scan_timestamp": {"$gte": today}})),
        ("Total pemindaian", await db.activities.count_documents({})),
        ("Menunggu persetujuan", await db.activities.count_documents({"status": "submitted"})),
        ("Jumlah lokasi", await db.locations.count_documents({})),
        ("Operator aktif", await db.users.count_documents({"role": "operator", "status": "active"})),
    ]:
        lines.append(f"- {label}: {count}")

    by_cat = await db.activities.aggregate([
        {"$match": {"scan_timestamp": {"$gte": today}}},
        {"$group": {"_id": "$activity_category_label", "count": {"$sum": 1}}},
    ]).to_list(20)
    if by_cat:
        lines.append("Pemindaian hari ini per kategori: " + ", ".join(f"{c['_id'] or 'Lainnya'} ({c['count']})" for c in by_cat))
    by_loc = await db.activities.aggregate([
        {"$match": {"scan_timestamp": {"$gte": today}}},
        {"$group": {"_id": "$scan_location", "count": {"$sum": 1}}},
    ]).to_list(20)
    if by_loc:
        lines.append("Pemindaian hari ini per lokasi: " + ", ".join(f"{l['_id'] or '-'} ({l['count']})" for l in by_loc))

    locs = await db.locations.find({}, {"_id": 0, "location_name": 1, "location_type": 1, "gps_coordinates": 1}).to_list(100)
    lines.append("Daftar lokasi: " + "; ".join(
        f"{l['location_name']} [{l.get('location_type')}]" + (f" GPS {l['gps_coordinates']}" if l.get("gps_coordinates") else "")
        for l in locs))

    recent = await db.activities.find({}, {"_id": 0}).sort("scan_timestamp", -1).to_list(15)
    lines.append("15 aktivitas terbaru:")
    for a in recent:
        lines.append(
            f"  - {a.get('scan_timestamp','')[:16]} | {a.get('inmate_name')} ({a.get('inmate_reg')}) | "
            f"{a.get('scan_location') or '-'} | {a.get('activity_category_label') or '-'} | "
            f"kondisi: {a.get('inmate_condition')} | status: {a.get('status')} | operator: {a.get('operator_name')}")

    alerts = await db.inmates.find({"medical_alert": {"$nin": [None, ""]}, "status": "active"},
                                   {"_id": 0, "full_name": 1, "registration_number": 1, "medical_alert": 1}).to_list(50)
    if alerts:
        lines.append("Peringatan medis aktif: " + "; ".join(
            f"{a['full_name']} ({a['registration_number']}): {a['medical_alert']}" for a in alerts))
    return "\n".join(lines)


class AiChatPayload(BaseModel):
    message: str
    session_id: Optional[str] = None


def sse(obj: dict) -> str:
    return f"data: {json.dumps(obj, ensure_ascii=False)}\n\n"


@api_router.post("/ai/chat")
async def ai_chat(payload: AiChatPayload, request: Request,
                  user: dict = Depends(require_roles("admin", "supervisor"))):
    if not EMERGENT_LLM_KEY:
        raise HTTPException(status_code=503, detail="Layanan AI belum dikonfigurasi")
    if not payload.message.strip():
        raise HTTPException(status_code=400, detail="Pesan tidak boleh kosong")
    sid = payload.session_id or new_id()
    await db.ai_messages.insert_one({
        "id": new_id(), "session_id": sid, "user_id": user["id"], "role": "user",
        "content": payload.message, "timestamp": now_iso(),
    })
    context = await build_ai_context()
    system = f"{AI_SYSTEM}\n\nDATA SISTEM SAAT INI:\n{context}"

    async def gen():
        full = []
        try:
            chat = LlmChat(api_key=EMERGENT_LLM_KEY, session_id=sid,
                           system_message=system).with_model(CLAUDE_PROVIDER, CLAUDE_MODEL)
            async for ev in chat.stream_message(UserMessage(text=payload.message)):
                if isinstance(ev, TextDelta):
                    full.append(ev.content)
                    yield sse({"text": ev.content})
                elif isinstance(ev, StreamDone):
                    break
        except Exception as e:
            logger.error(f"AI chat error: {e}")
            yield sse({"error": "Layanan AI sedang bermasalah. Coba lagi."})
        answer = "".join(full)
        if answer:
            await db.ai_messages.insert_one({
                "id": new_id(), "session_id": sid, "user_id": user["id"], "role": "assistant",
                "content": answer, "timestamp": now_iso(),
            })
        yield sse({"done": True, "session_id": sid})

    return StreamingResponse(gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@api_router.get("/ai/sessions")
async def ai_sessions(user: dict = Depends(require_roles("admin", "supervisor"))):
    pipeline = [
        {"$match": {"user_id": user["id"]}},
        {"$sort": {"timestamp": 1}},
        {"$group": {"_id": "$session_id", "title": {"$first": "$content"}, "last": {"$last": "$timestamp"}}},
        {"$sort": {"last": -1}},
        {"$limit": 20},
    ]
    items = await db.ai_messages.aggregate(pipeline).to_list(20)
    return [{"session_id": i["_id"], "title": (i["title"] or "")[:60], "last": i["last"]} for i in items]


@api_router.get("/ai/sessions/{session_id}/messages")
async def ai_session_messages(session_id: str, user: dict = Depends(require_roles("admin", "supervisor"))):
    return await db.ai_messages.find(
        {"session_id": session_id, "user_id": user["id"]},
        {"_id": 0, "role": 1, "content": 1, "timestamp": 1},
    ).sort("timestamp", 1).to_list(200)


class AiReportPayload(BaseModel):
    period: str = "today"
    officer_name: Optional[str] = None
    officer_nip: Optional[str] = None
    recipient: Optional[str] = None
    place: Optional[str] = None
    location_id: Optional[str] = None
    category: Optional[str] = None


@api_router.post("/ai/report")
async def ai_report(payload: AiReportPayload, request: Request,
                    user: dict = Depends(require_roles("admin", "supervisor"))):
    if not EMERGENT_LLM_KEY:
        raise HTTPException(status_code=503, detail="Layanan AI belum dikonfigurasi")
    now = datetime.now(timezone.utc)
    if payload.period == "week":
        start = (now - timedelta(days=7)).date().isoformat()
        period_label = "7 hari terakhir"
    else:
        start = now.date().isoformat()
        period_label = "hari ini"
    settings = await get_settings_doc()
    q = {"scan_timestamp": {"$gte": start}}
    scope_parts = []
    if payload.location_id:
        loc = await db.locations.find_one({"id": payload.location_id})
        if loc:
            q["location_id"] = payload.location_id
            scope_parts.append(f"lokasi {loc['location_name']}")
    if payload.category:
        cat_label = next((c["label"] for c in settings["activity_categories"]
                          if c["key"] == payload.category), payload.category)
        q["activity_category"] = payload.category
        scope_parts.append(f"kategori {cat_label}")
    if scope_parts:
        period_label += " khusus " + " dan ".join(scope_parts)
    else:
        period_label += " (seluruh lokasi dan kategori)"
    items = await db.activities.find(q, {"_id": 0}).sort("scan_timestamp", 1).to_list(2000)
    rows = "\n".join(
        f"- {a.get('scan_timestamp','')[:16]} | {a.get('inmate_name')} ({a.get('inmate_reg')}) | "
        f"{a.get('scan_location') or '-'} | {a.get('activity_category_label') or '-'} | "
        f"kondisi: {a.get('inmate_condition')} | status: {a.get('status')}"
        for a in items) or "(tidak ada aktivitas pada periode ini)"
    officer = (payload.officer_name or "").strip() or "____________________"
    nip = (payload.officer_nip or "").strip() or "____________________"
    recipient = (payload.recipient or "").strip() or f"Kepala {settings['institution_name']}"
    place = (payload.place or "").strip() or "________________"
    today_id = now.strftime("%d") + " " + [
        "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember",
    ][now.month] + f" {now.year}"
    prompt = (
        f"Buatkan LAPORAN ATENSI PIMPINAN resmi untuk {settings['institution_name']} berdasarkan data aktivitas "
        f"periode {period_label} berikut. Ikuti PERSIS struktur format di bawah ini (tanpa heading markdown '#', "
        "judul dan nomor bagian menggunakan huruf kapital):\n\n"
        "LAPORAN ATENSI PIMPINAN\n\n"
        f"Kepada:\nYth. {recipient}\n\n"
        f"Dari:\n{officer}\n\n"
        "I. PERISTIWA/KEGIATAN\n- (ringkasan singkat jenis kegiatan/insiden dari data aktivitas)\n\n"
        "II. URAIAN/KEGIATAN\n- (uraian kronologis tiap kegiatan penting: sebutkan hari/tanggal, pukul, nama warga binaan, "
        "lokasi/blok, petugas/operator yang mencatat, kategori pembinaan, dan kondisi warga binaan; bila ada warga binaan "
        "berkondisi sakit/perlu perhatian atau aktivitas yang ditolak supervisor, wajib diuraikan; akhiri dengan kalimat "
        "'Situasi aman dan kondusif.' bila tidak ada insiden keamanan)\n\n"
        "III. TEMPAT KEGIATAN\n- (daftar lokasi kegiatan dari data, sertakan nama institusi)\n\n"
        "IV. FOTO/DOKUMENTASI\n- (tulis 'Terlampir' jika ada dokumentasi pada data, jika tidak: 'Tidak ada')\n\n"
        "V. TINDAK LANJUT\n- Melaporkan kepada pimpinan. (tambahkan rekomendasi tindak lanjut bila ada temuan medis "
        "atau aktivitas bermasalah)\n\n"
        "VI. PENUTUP\n- Demikian Laporan Atensi ini dibuat. Selanjutnya mohon arahan dan petunjuk, terima kasih.\n\n"
        f"{place}, {today_id}\nAnggota jaga\n\n\nTTD\n\n\n{officer}\nNIP. {nip}\n\n"
        f"DATA AKTIVITAS ({len(items)} entri):\n{rows}"
    )

    async def gen():
        try:
            chat = LlmChat(api_key=EMERGENT_LLM_KEY, session_id=f"report-{new_id()}",
                           system_message=AI_SYSTEM).with_model(CLAUDE_PROVIDER, CLAUDE_MODEL)
            async for ev in chat.stream_message(UserMessage(text=prompt)):
                if isinstance(ev, TextDelta):
                    yield sse({"text": ev.content})
                elif isinstance(ev, StreamDone):
                    break
        except Exception as e:
            logger.error(f"AI report error: {e}")
            yield sse({"error": "Layanan AI sedang bermasalah. Coba lagi."})
        yield sse({"done": True})

    await log_audit(user, "activities", "-", "export",
                    {"type": "ai_report", "period": payload.period,
                     "location_id": payload.location_id, "category": payload.category}, request)
    return StreamingResponse(gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


async def seed():
    await db.users.create_index("username", unique=True)
    await db.inmates.create_index("registration_number", unique=True)
    await db.activities.create_index("inmate_id")
    await db.activities.create_index("status")
    await db.activities.create_index("scan_timestamp")
    await db.audit_log.create_index("timestamp")
    await db.security_crossings.create_index("inmate_id")
    await db.security_crossings.create_index("scan_timestamp")

    await get_settings_doc()
    await migrate_settings_modules()

    admin_username = os.environ.get("ADMIN_USERNAME", "admin")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@123")
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com")
    existing = await db.users.find_one({"username": admin_username})
    if not existing:
        await db.users.insert_one({
            "id": new_id(), "username": admin_username, "email": admin_email,
            "phone": None, "full_name": "Administrator", "position": "Kepala Administrasi",
            "assigned_location": None, "device_name": None, "role": "admin",
            "hashed_password": hash_password(admin_password), "status": "active",
            "start_date": None, "end_date": None, "last_login": None, "last_activity": None,
            "created_at": now_iso(), "updated_at": now_iso(),
        })
        logger.info("Admin user seeded")
    elif not verify_password(admin_password, existing["hashed_password"]):
        await db.users.update_one({"username": admin_username},
                                  {"$set": {"hashed_password": hash_password(admin_password),
                                            "email": admin_email}})

    if not await db.locations.find_one({}):
        locs = [
            ("Masjid At-Taubah", "religious"), ("Bengkel Kerja", "skills"),
            ("Poliklinik", "health"), ("Ruang Kelas", "education"),
            ("Lapangan Olahraga", "sports"), ("Pos Keamanan Blok A", "security"),
        ]
        for name, ltype in locs:
            await db.locations.insert_one({
                "id": new_id(), "location_name": name, "location_type": ltype,
                "gps_coordinates": None, "description": None, "created_at": now_iso(),
            })
        logger.info("Default locations seeded")

    if not await db.inmates.find_one({}):
        samples = [
            ("WB-2024-0001", "Budi Santoso", "A", "Narkotika", None),
            ("WB-2024-0002", "Andi Wijaya", "B", "Pencurian", "Hipertensi - kontrol rutin"),
            ("WB-2024-0003", "Slamet Riyadi", "A", "Penggelapan", None),
        ]
        for reg, name, block, crime, med in samples:
            await db.inmates.insert_one({
                "id": new_id(), "registration_number": reg, "full_name": name,
                "identity_number": None, "photo_url": None, "status": "active",
                "date_entry": "2024-01-15", "estimated_release_date": "2027-01-15",
                "cell_block": block, "crime_category": crime, "medical_alert": med,
                "barcode_data": reg, "created_at": now_iso(), "updated_at": now_iso(),
            })
        logger.info("Sample inmates seeded")

    if not await db.users.find_one({"username": "supervisor"}):
        await db.users.insert_one({
            "id": new_id(), "username": "supervisor", "email": None, "phone": None,
            "full_name": "Supervisor Pembinaan", "position": "Kasi Pembinaan",
            "assigned_location": None, "device_name": None, "role": "supervisor",
            "hashed_password": hash_password("Supervisor@123"), "status": "active",
            "start_date": None, "end_date": None, "last_login": None, "last_activity": None,
            "created_at": now_iso(), "updated_at": now_iso(),
        })
    if not await db.users.find_one({"username": "operator1"}):
        loc = await db.locations.find_one({"location_name": "Masjid At-Taubah"})
        await db.users.insert_one({
            "id": new_id(), "username": "operator1", "email": None, "phone": None,
            "full_name": "Operator Masjid", "position": "Petugas Scan",
            "assigned_location": loc["id"] if loc else None,
            "device_name": "Tablet Pos Masjid", "role": "operator",
            "hashed_password": hash_password("Operator@123"), "status": "active",
            "start_date": None, "end_date": None, "last_login": None, "last_activity": None,
            "created_at": now_iso(), "updated_at": now_iso(),
        })


@app.on_event("startup")
async def startup():
    await seed()


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

